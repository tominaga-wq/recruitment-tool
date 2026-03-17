import streamlit as st
import openpyxl
import anthropic
import json
import io
from google import genai
from google.genai import types
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError

import os
import time
import datetime

CLAUDE_API_KEY = st.secrets["CLAUDE_API_KEY"]
GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
APP_PASSWORD = st.secrets.get("APP_PASSWORD", "Unipo1113!")
SLACK_BOT_TOKEN = st.secrets.get("SLACK_BOT_TOKEN", "")
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "Lionking7")

# Excelファイルパス（ローカル or アップロードファイル）
EXCEL_FILENAME = "求人確度出力マスターシート.xlsx"
LOCAL_EXCEL_PATH = r"C:\Users\takeu\Downloads\求人確度出力マスターシート.xlsx"
EXCEL_PATH = LOCAL_EXCEL_PATH if os.path.exists(LOCAL_EXCEL_PATH) else EXCEL_FILENAME

st.set_page_config(page_title="求人マッチングツール", page_icon="🎯", layout="wide")

# ---- パスワード認証 ----
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

if not st.session_state.authenticated:
    st.title("🔐 求人マッチングツール")
    pw = st.text_input("パスワードを入力してください", type="password")
    if st.button("ログイン"):
        if pw == ADMIN_PASSWORD:
            st.session_state.authenticated = True
            st.session_state.is_admin = True
            st.rerun()
        elif pw == APP_PASSWORD:
            st.session_state.authenticated = True
            st.session_state.is_admin = False
            st.rerun()
        else:
            st.error("パスワードが違います")
    st.stop()


def extract_file_text(uploaded_file) -> str:
    name = uploaded_file.name.lower()
    if name.endswith(".txt"):
        return uploaded_file.read().decode("utf-8", errors="replace")
    elif name.endswith(".pdf"):
        import pdfplumber
        text = ""
        with pdfplumber.open(io.BytesIO(uploaded_file.read())) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
        return text
    elif name.endswith(".docx"):
        from docx import Document
        doc = Document(io.BytesIO(uploaded_file.read()))
        return "\n".join([p.text for p in doc.paragraphs])
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()))
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                line = "　".join([str(c) for c in row if c is not None])
                if line.strip():
                    text += line + "\n"
        return text
    else:
        return uploaded_file.read().decode("utf-8", errors="replace")


@st.cache_data
def load_company_requirements():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    companies = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "候補者一覧":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        info = {"company_name": "", "position": "", "must": "", "want": "", "ng": "", "description": ""}
        for row in rows:
            if not row[0]:
                continue
            label = str(row[0])
            value = str(row[1]) if row[1] else ""
            if "求人データ" in label:
                parts = label.split("▶▶")
                info["company_name"] = parts[-1].strip() if len(parts) > 1 else label.replace("求人データ", "").strip()
            elif "参考ポジション" in label:
                info["position"] = label
            elif "必須" in label or "Must" in label:
                info["must"] = value
            elif "歓迎" in label or "Want" in label:
                info["want"] = value
            elif "NG条件" in label or "NG" in label:
                info["ng"] = value
            elif "業務内容" in label or "仕事内容" in label:
                info["description"] = value
        if not info["company_name"]:
            parts = sheet_name.split("_", 1)
            info["company_name"] = parts[1] if len(parts) > 1 else sheet_name
        companies[sheet_name] = info
    return companies


@st.cache_data
def load_candidates():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["候補者一覧"]
    candidates = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0]:
            candidates.append({
                "name": str(row[0] or ""),
                "decided_company": str(row[1] or ""),
                "job_type": str(row[2] or ""),
                "skills": str(row[3] or ""),
                "age": str(row[4] or ""),
                "job_change_count": str(row[5] or 0),
                "recent_industry": str(row[6] or ""),
                "total_experience": str(row[7] or ""),
                "summary": str(row[8] or ""),
                "comment": str(row[9] or ""),
            })
    return candidates


def build_hire_profiles(candidates: list, companies: dict) -> str:
    """各企業の過去内定者プロフィールをまとめたテキストを生成"""
    # 企業名→シート名のマッピング
    company_name_to_sheet = {info["company_name"]: sheet for sheet, info in companies.items()}

    # 内定先企業ごとに候補者をグループ化
    profiles_by_company = {}
    for c in candidates:
        decided = c["decided_company"].strip()
        if not decided or decided == "None":
            continue
        # 部分一致でシート名を探す
        matched_sheet = None
        for company_name, sheet in company_name_to_sheet.items():
            if decided in company_name or company_name in decided:
                matched_sheet = sheet
                break
        if matched_sheet:
            if matched_sheet not in profiles_by_company:
                profiles_by_company[matched_sheet] = []
            profiles_by_company[matched_sheet].append(c)

    if not profiles_by_company:
        return ""

    text = "## 各企業の過去内定者プロフィール（実績データ）\n"
    for sheet, hires in profiles_by_company.items():
        company_name = companies[sheet]["company_name"]
        text += f"\n### {company_name}（{len(hires)}名が内定）\n"
        for h in hires:
            text += f"- {h['job_type']}／{h['recent_industry']}／経験{h['total_experience']}／年齢{h['age']}歳／転職{h['job_change_count']}回\n"
    return text


def build_company_list(companies: dict) -> str:
    text = ""
    for sheet_name, info in companies.items():
        ng_line = f"\n【NG要件】{info['ng']}" if info.get('ng') else ""
        text += f"""
---
【企業シート名】{sheet_name}
【企業名】{info['company_name']}
【ポジション】{info['position']}
【必須要件（Must）】{info['must']}
【歓迎要件（Want）】{info['want']}{ng_line}
【業務内容】{info['description']}
"""
    return text


# ---- STEP 1: Claudeで上位8社を選出 ----
def step1_rank_companies(candidate_text: str, companies: dict, hire_profiles: str = "") -> list:
    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    company_list = build_company_list(companies)

    prompt = f"""あなたは人材紹介会社のシニアキャリアアドバイザーです。

求職者情報と求人要件を照合し、成約率が最も高くなるよう上位8社を選んでください。
完全にマッチしない場合でも、相対的に可能性が高い順に必ず8社を選出してください。
また、求職者の転職理由・今回の転職で目指しているキャリア・将来プランを文脈から推測してください。

スコアリングの際は、求人要件とのマッチに加えて、各企業の過去内定者プロフィールとの類似度も重視してください。
過去内定者に似た特徴（職種・業界・経験年数・年齢・スキル）を持つ候補者は内定確度が高いと判断してください。

【重要】各企業に【NG要件】が設定されている場合、候補者がそのNG要件に1つでも該当するならその企業は必ず選出から除外してください。NG要件に該当する企業は8社のカウントに含めないでください。

## 各項目を1〜5点で採点してください
- S（スキルマッチ）：技術・経験の合致度。5=完全一致、1=ほぼ合致なし
- C（キャリアの連続性）：職種・業界の親和性。5=自然なステップアップ、1=全く異なる
- A（志向性の一致）：本人の希望条件・転職理由との合致度。5=まさに求めている、1=全く合わない
- H（採用ハードル）：選考難易度。5=最難関（書類落ちリスク大）、1=容易（ほぼ確実に通過）

【採点の注意・必須】H（採用ハードル）は8社の中で必ず以下の分布にしてください：
- H=4または5：必ず2〜3社（難関企業・競争率が高い企業）
- H=3：2〜3社（標準的な難易度）
- H=1または2：必ず2〜3社（通過率が高い・条件が緩い企業）
全社をH=3にすることは厳禁です。実際の採用難易度に基づいて正直に採点してください。

## 求職者情報
{candidate_text}

## 各企業の求人要件
{company_list}

{hire_profiles}

## 出力形式（JSONのみ返してください）
```json
{{
  "candidate_summary": {{
    "name": "氏名（不明なら「不明」）",
    "inferred_reason": "推測した転職理由（2〜3文）",
    "inferred_career": "推測した今回の転職で目指しているキャリア（2〜3文）",
    "inferred_future": "推測した将来プラン（2〜3文）"
  }},
  "top8": [
    {{
      "rank": 1,
      "sheet_name": "シート名",
      "company_name": "企業名",
      "position": "ポジション名",
      "S": 4,
      "C": 3,
      "A": 5,
      "H": 4,
      "match_reason": "スキル・経験面からのマッチ理由（候補者の具体的な実績・数字を引用して3〜4文）"
    }}
  ]
}}
```
"""
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = message.content[0].text
    try:
        start = raw.find("{")
        end = raw.rfind("}") + 1
        return json.loads(raw[start:end])
    except Exception:
        return {}


# ---- STEP 2: Geminiでリアルタイム企業情報を収集 ----
def step2_search_companies(company_names: list) -> str:
    client = genai.Client(api_key=GEMINI_API_KEY)
    names_str = "、".join(company_names)

    query = f"""以下の企業について、採用担当者が求職者に訴求するために役立つ情報を各社200文字以上で調査してください。

対象企業：{names_str}

各社について以下を100文字程度で簡潔にまとめてください：
- 事業内容・強み
- 成長性・社風
- 求職者へのキャリアメリット

企業名を見出しにして、各社の情報をまとめてください。"""

    # Google検索グラウンディングで試み、失敗したら検索なしにフォールバック
    for model_name in ["gemini-2.0-flash", "gemini-2.0-flash-lite", "gemini-2.5-flash"]:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=query,
                config=types.GenerateContentConfig(
                    tools=[types.Tool(google_search=types.GoogleSearch())]
                ),
            )
            return response.text
        except Exception:
            pass

    # 検索なし（学習データから生成）
    for model_name in ["gemini-2.0-flash-lite", "gemini-2.0-flash", "gemini-2.5-flash"]:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=query,
            )
            return response.text
        except Exception:
            continue

    raise Exception("利用可能なGeminiモデルが見つかりませんでした")


# ---- STEP 3: Geminiで訴求文を肉付け生成 ----
def step3_enrich_pitches(candidate_text: str, step1_data: dict, gemini_info: str, companies: dict) -> dict:
    client = genai.Client(api_key=GEMINI_API_KEY)

    top8 = step1_data.get("top8", [])
    summary = step1_data.get("candidate_summary", {})

    # 上位8社の求人要件だけ抽出
    top_requirements = ""
    for item in top8:
        sheet = item.get("sheet_name", "")
        if sheet in companies:
            info = companies[sheet]
            top_requirements += f"""
---
【企業名】{item['company_name']}（rank {item['rank']}）
【ポジション】{item['position']}
【必須要件】{info['must']}
【歓迎要件】{info['want']}
【業務内容】{info['description']}
"""

    prompt = f"""あなたは人材紹介会社のシニアキャリアアドバイザーです。

以下の情報をもとに、上位8社それぞれについて求職者への訴求文を生成してください。
各訴求文は**必ず200文字以上**で、Geminiの企業リサーチ情報も積極的に活用して具体的・肉厚に書いてください。

## 求職者情報
{candidate_text}

## AIが推測した転職背景
- 転職理由：{summary.get('inferred_reason', '')}
- 今回の転職で目指しているキャリア：{summary.get('inferred_career', '')}
- 将来プラン：{summary.get('inferred_future', '')}

## 上位8社の求人要件
{top_requirements}

## Geminiが収集した企業リサーチ情報（Google検索結果）
{gemini_info}

## 出力形式（JSONのみ返してください。余分なテキストは不要です）
[
  {{
    "rank": 1,
    "sheet_name": "シート名",
    "company_name": "企業名",
    "position": "ポジション名",
    "match_score": 85,
    "match_reason": "スキル・経験面からのマッチ理由（候補者の具体的な実績・数字を引用して3〜4文）",
    "pitch_reason": "【転職理由との合致】候補者の転職理由とこの求人がどう合致するか。Geminiの企業情報も交えて200文字以上で具体的に",
    "pitch_career": "【今回の転職で目指しているキャリアとの合致】候補者が目指すキャリアに対してこの求人でどんな経験・成長が得られるか。Geminiの企業情報も交えて200文字以上で具体的に",
    "pitch_future": "【将来プランとの合致】候補者の長期ビジョンに対してこの企業・ポジションがどうつながるか。Geminiの企業情報も交えて200文字以上で具体的に"
  }}
]
"""
    raw = ""
    for model_name in ["gemini-2.0-flash", "gemini-2.0-flash-lite", "gemini-2.5-flash"]:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=prompt,
            )
            raw = response.text
            break
        except Exception:
            continue

    if not raw:
        st.error("Step3: Geminiによる訴求文生成に失敗しました")
        return {}

    try:
        start = raw.find("[")
        end = raw.rfind("]") + 1
        if start == -1 or end == 0:
            st.error("Step3: JSONが見つかりませんでした")
            st.code(raw[:2000])
            return {}
        results = json.loads(raw[start:end])
        score_map = {item["sheet_name"]: item for item in top8}
        for r in results:
            sn = r.get("sheet_name", "")
            if sn in score_map and "match_score" not in r:
                r["match_score"] = score_map[sn].get("match_score", 0)
        return {"candidate_summary": summary, "results": results}
    except Exception as e:
        st.error(f"Step3 解析エラー: {e}")
        st.code(raw[:3000])
        return {}


# 検証用: TrueにするとStep1のマッチング結果のみ表示（訴求文生成をスキップ）
FAST_MODE = True


def run_analysis(candidate_text: str, companies: dict):
    candidates = load_candidates()
    hire_profiles = build_hire_profiles(candidates, companies)

    if FAST_MODE:
        progress = st.progress(0, text="Claudeがマッチング中...")
        step1_data = step1_rank_companies(candidate_text, companies, hire_profiles)
        progress.progress(100, text="完了！")
        progress.empty()
        if not step1_data:
            st.error("Step1の分析に失敗しました")
            return
        show_results_fast(step1_data)
        return

    progress = st.progress(0, text="Step 1/3 : Claudeがマッチング中...")
    step1_data = step1_rank_companies(candidate_text, companies, hire_profiles)
    if not step1_data:
        st.error("Step1の分析に失敗しました")
        return

    top8 = step1_data.get("top8", [])
    company_names = [item["company_name"] for item in top8]

    progress.progress(33, text="Step 2/3 : GeminiがGoogle検索で企業情報を収集中...")

    try:
        gemini_info = step2_search_companies(company_names)
    except Exception as e:
        st.warning(f"Gemini検索でエラーが発生しました（スキップして続行）: {e}")
        gemini_info = "企業情報の取得に失敗しました。"

    progress.progress(66, text="Step 3/3 : Geminiが訴求文を生成中...")

    final_data = step3_enrich_pitches(candidate_text, step1_data, gemini_info, companies)

    progress.progress(100, text="完了！")
    progress.empty()

    if final_data:
        show_results(final_data)
    else:
        st.error("最終出力の生成に失敗しました")


def classify_company(r: dict) -> str:
    """S/C/A/Hスコアから チャレンジ/本命/セーフティー を判定"""
    S = r.get("S", 0)
    C = r.get("C", 0)
    A = r.get("A", 0)
    H = r.get("H", 0)
    # チャレンジ：難易度が高く、志向性も最低限ある
    if H >= 4 and A >= 3:
        return "チャレンジ"
    # セーフティー：難易度が低くスキルは十分
    if H <= 2 and A >= 2:
        return "セーフティー"
    # 本命：志向性・実力ともに一定以上
    if A >= 3 and (S + C) >= 5:
        return "本命"
    return "参考"


def show_results_fast(data: dict):
    summary = data.get("candidate_summary", {})
    top8 = data.get("top8", [])
    name = summary.get("name", "候補者")

    st.success(f"✅ 分析完了！ {name}さんへの推奨求人")

    with st.expander("📌 AIが推測した転職背景", expanded=True):
        col1, col2, col3 = st.columns(3)
        col1.markdown(f"**転職理由**\n\n{summary.get('inferred_reason', '')}")
        col2.markdown(f"**今回の転職で目指しているキャリア**\n\n{summary.get('inferred_career', '')}")
        col3.markdown(f"**将来プラン**\n\n{summary.get('inferred_future', '')}")

    # 分類
    for r in top8:
        r["_category"] = classify_company(r)

    categories = [
        ("本命", "⭐ 本命"),
        ("チャレンジ", "🔥 チャレンジ"),
        ("セーフティー", "🛡️ セーフティー"),
        ("参考", "📋 参考"),
    ]

    for category, label in categories:
        items = [r for r in top8 if r["_category"] == category]
        if not items:
            continue
        st.markdown(f"### {label}")
        for r in items:
            S, C, A, H = r.get("S", "-"), r.get("C", "-"), r.get("A", "-"), r.get("H", "-")
            header = f"**{r['company_name']}**　S:{S} C:{C} A:{A} H:{H}"
            with st.expander(header, expanded=True):
                col_left, col_right = st.columns([3, 1])
                with col_left:
                    st.write(f"**ポジション：** {r.get('position', '')}")
                    st.markdown(f"**マッチ理由**\n\n{r.get('match_reason', '')}")
                with col_right:
                    st.metric("S スキル", f"{S}/5")
                    st.metric("C 連続性", f"{C}/5")
                    st.metric("A 志向性", f"{A}/5")
                    st.metric("H ハードル", f"{H}/5")
        st.markdown("---")


def show_results(data: dict):
    summary = data.get("candidate_summary", {})
    results = data.get("results", [])
    name = summary.get("name", "候補者")

    st.success(f"✅ 分析完了！ {name}さんへの推奨求人 上位8社")

    with st.expander("📌 AIが推測した転職背景", expanded=True):
        col1, col2, col3 = st.columns(3)
        col1.markdown(f"**転職理由**\n\n{summary.get('inferred_reason', '')}")
        col2.markdown(f"**今回の転職で目指しているキャリア**\n\n{summary.get('inferred_career', '')}")
        col3.markdown(f"**将来プラン**\n\n{summary.get('inferred_future', '')}")

    st.markdown("---")

    for r in results:
        with st.expander(f"**#{r['rank']} {r['company_name']}**　スコア: {r.get('match_score', '-')}/100", expanded=r['rank'] <= 3):
            col_left, col_right = st.columns([3, 1])
            with col_left:
                st.write(f"**ポジション：** {r.get('position', '')}")
                st.markdown(f"**マッチ理由（スキル・経験）**\n\n{r.get('match_reason', '')}")
                st.markdown("---")
                st.markdown(f"**転職理由との合致**\n\n{r.get('pitch_reason', '')}")
                st.markdown(f"**今回の転職で目指しているキャリアとの合致**\n\n{r.get('pitch_career', '')}")
                st.markdown(f"**将来プランとの合致**\n\n{r.get('pitch_future', '')}")
            with col_right:
                st.metric("内定確度スコア", f"{r.get('match_score', '-')}/100")


# ---- Slack連携: チャンネル一覧取得 ----
def get_slack_channels() -> list:
    client = WebClient(token=SLACK_BOT_TOKEN)
    channels = []
    try:
        res = client.conversations_list(types="public_channel,private_channel", limit=200)
        for ch in res["channels"]:
            channels.append({"id": ch["id"], "name": ch["name"]})
    except SlackApiError:
        pass
    return channels


# ---- Slack連携: メッセージ取得（スレッド含む・日数指定） ----
def get_slack_messages(channel_id: str, days: int = 30) -> str:
    client = WebClient(token=SLACK_BOT_TOKEN)
    all_text = ""
    oldest = str((datetime.datetime.utcnow() - datetime.timedelta(days=days)).timestamp())
    cursor = None
    try:
        while True:
            kwargs = {"channel": channel_id, "oldest": oldest, "limit": 200}
            if cursor:
                kwargs["cursor"] = cursor
            res = client.conversations_history(**kwargs)
            for msg in reversed(res["messages"]):
                text = msg.get("text", "")
                ts = msg.get("ts", "")
                all_text += f"{text}\n"
                if msg.get("reply_count", 0) > 0:
                    thread_res = client.conversations_replies(channel=channel_id, ts=ts)
                    for reply in thread_res["messages"][1:]:
                        all_text += f"  └ {reply.get('text', '')}\n"
            if res.get("has_more") and res.get("response_metadata", {}).get("next_cursor"):
                cursor = res["response_metadata"]["next_cursor"]
            else:
                break
    except SlackApiError as e:
        st.error(f"Slack取得エラー: {e}")
    except Exception as e:
        st.error(f"予期せぬエラー: {e}")
    return all_text


# ---- Slack連携: Claudeで求人要件を抽出 ----
def extract_requirements_from_slack(slack_text: str, companies: dict) -> list:
    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    # シート名と企業名の両方を渡す
    company_list_str = "\n".join(
        f"- {sheet}（{info['company_name']}）" for sheet, info in companies.items()
    )
    # Slackテキストが長すぎる場合は先頭15000文字に絞る
    slack_text_trimmed = slack_text[:15000] if len(slack_text) > 15000 else slack_text

    prompt = f"""以下のSlackメッセージから、各企業の求人要件を「必須要件」と「歓迎要件」に分けて抽出してください。

## 分類の基準
- **必須要件（must）**：「必須」「必要」「要件」「書類要件」「〜であること」「〜経験必須」「〜以上」など、応募に必要な条件として書かれているもの
- **歓迎要件（want）**：「歓迎」「尚可」「あれば」「望ましい」「プラス」「あると嬉しい」など、あれば有利という表現のもの

業務内容は抽出不要です。要件のみ抽出してください。

## 登録済み企業一覧（シート名＋日本語名）
{company_list_str}

## Slackメッセージ
{slack_text_trimmed}

## マッチングルール
- Slackの企業名と登録済み企業一覧は表記が異なります（例：「XMile社」→「エックスマイル」、「リクルート」→「リクルートスタッフィング」など）
- 部分一致・読み仮名・略称でも積極的にマッチングしてください
- company_nameには登録済み企業一覧の**日本語名**をそのまま使用してください

## 出力形式（JSONのみ返してください）
[
  {{
    "company_name": "登録済み企業一覧の日本語名をそのまま記載",
    "must": "抽出した必須要件（なければ空文字）",
    "want": "抽出した歓迎要件（なければ空文字）"
  }}
]

企業名が一覧にどうしても一致しない場合はスキップしてください。"""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = message.content[0].text
    try:
        start = raw.find("[")
        end = raw.rfind("]") + 1
        return json.loads(raw[start:end])
    except Exception:
        return []


# ---- Slack連携: Excelを更新してダウンロード用バイトを返す ----
def update_excel_with_requirements(updates: list, companies: dict) -> bytes:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    # 企業名→シート名マッピング
    name_to_sheet = {info["company_name"]: sheet for sheet, info in companies.items()}

    for update in updates:
        company_name = update.get("company_name", "")
        sheet_name = name_to_sheet.get(company_name)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            if not row[0].value:
                continue
            label = str(row[0].value)
            if ("必須" in label or "Must" in label) and update.get("must"):
                # 必須要件は上書き
                row[1].value = update["must"]
            elif ("歓迎" in label or "Want" in label) and update.get("want"):
                # 歓迎要件は追記
                existing = str(row[1].value) if row[1].value else ""
                new_want = update["want"]
                if existing and new_want not in existing:
                    row[1].value = existing + "\n" + new_want
                elif not existing:
                    row[1].value = new_want

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---- UI ----
st.title("🎯 求人マッチングツール")
st.caption("候補者情報を貼り付け or ファイルアップロードするだけで、内定確度の高い求人8社＋訴求方法を自動生成します")

companies = load_company_requirements()
candidates = load_candidates()

if st.session_state.is_admin:
    tab1, tab2, tab3 = st.tabs(["📋 既存候補者から選ぶ", "📄 テキスト貼り付け / ファイルアップロード", "💬 Slack連携で求人要件を更新"])
else:
    tab1, tab2 = st.tabs(["📋 既存候補者から選ぶ", "📄 テキスト貼り付け / ファイルアップロード"])
    tab3 = None

# ---- Tab1 ----
with tab1:
    candidate_names = [f"{c['name']}（{c['job_type']}）" for c in candidates]
    selected = st.selectbox("候補者を選択", ["--- 選択してください ---"] + candidate_names)

    if selected != "--- 選択してください ---":
        idx = candidate_names.index(selected)
        c = candidates[idx]
        with st.expander("候補者情報を確認", expanded=True):
            col1, col2, col3 = st.columns(3)
            col1.metric("年齢", f"{c['age']}歳")
            col2.metric("転職回数", f"{c['job_change_count']}回")
            col3.metric("経験年数", c['total_experience'])
            st.write(f"**経験職種：** {c['job_type']}")
            st.write(f"**主要スキル：** {c['skills']}")
            st.write(f"**直近業種：** {c['recent_industry']}")
            st.write(f"**職務要約：** {c['summary']}")

        if st.button("🔍 マッチング開始", type="primary", key="btn1"):
            candidate_text = f"""氏名：{c['name']}
年齢：{c['age']}歳 / 転職回数：{c['job_change_count']}回
経験職種：{c['job_type']}
主要スキル：{c['skills']}
直近業種：{c['recent_industry']} / 合計経験年数：{c['total_experience']}
職務要約・強み：{c['summary']}
評価コメント：{c['comment']}"""
            run_analysis(candidate_text, companies)

# ---- Tab2 ----
with tab2:
    st.markdown("職務経歴書・面談メモ・候補者プロフィールなど、**何でも貼り付けかアップロード**してください。両方同時もOKです。")

    col1, col2 = st.columns([3, 2])
    with col1:
        pasted_text = st.text_area(
            "テキストを貼り付け",
            placeholder="職務経歴書、面談メモ、スカウト文面など何でもOK",
            height=300,
        )
    with col2:
        uploaded_files = st.file_uploader(
            "ファイルをアップロード（複数可）",
            type=["txt", "pdf", "docx", "xlsx"],
            accept_multiple_files=True,
        )

    if st.button("🔍 マッチング開始", type="primary", key="btn2"):
        combined_text = ""
        if pasted_text.strip():
            combined_text += "【貼り付けテキスト】\n" + pasted_text.strip() + "\n\n"
        if uploaded_files:
            for f in uploaded_files:
                file_text = extract_file_text(f)
                combined_text += f"【ファイル: {f.name}】\n" + file_text.strip() + "\n\n"

        if not combined_text.strip():
            st.warning("テキストを貼り付けるかファイルをアップロードしてください")
        else:
            run_analysis(combined_text, companies)

# ---- Tab3: Slack連携（管理者のみ） ----
if st.session_state.is_admin:
    with tab3:
        st.markdown("Slackチャンネルから求人要件を読み取り、Excelを更新してダウンロードできます。")

        if not SLACK_BOT_TOKEN:
            st.error("Slack Bot Tokenが設定されていません")
        else:
            channels = get_slack_channels()
            if not channels:
                st.error("チャンネルが取得できませんでした。Botをチャンネルに招待してください。")
            else:
                channel_options = {ch["name"]: ch["id"] for ch in channels}
                selected_channel = st.selectbox("チャンネルを選択", list(channel_options.keys()))
                days = st.slider("取得する期間（日数）", min_value=7, max_value=365, value=90, step=7)

                if st.button("📥 Slackから読み取る", type="primary"):
                    with st.spinner("Slackメッセージを取得中..."):
                        channel_id = channel_options[selected_channel]
                        slack_text = get_slack_messages(channel_id, days=days)

                    if not slack_text.strip():
                        st.warning("メッセージが取得できませんでした")
                    else:
                        with st.spinner("Claudeが求人要件を抽出中..."):
                            extracted = extract_requirements_from_slack(slack_text, companies)

                        if not extracted:
                            st.warning("求人要件を抽出できませんでした。チャンネルに要件に関する投稿があるか確認してください。")
                        else:
                            st.success(f"{len(extracted)}社分の要件を抽出しました。内容を確認してください。")
                            st.session_state["slack_extracted"] = extracted

                if "slack_extracted" in st.session_state and st.session_state["slack_extracted"]:
                    extracted = st.session_state["slack_extracted"]
                    for i, item in enumerate(extracted):
                        col_check, col_content = st.columns([1, 10])
                        with col_check:
                            checked = st.checkbox("更新する", value=True, key=f"check_{i}")
                            extracted[i]["_include"] = checked
                        with col_content:
                            with st.expander(f"**{item['company_name']}**", expanded=True):
                                extracted[i]["must"] = st.text_area("必須要件", value=item.get("must", ""), key=f"must_{i}")
                                extracted[i]["want"] = st.text_area("歓迎要件", value=item.get("want", ""), key=f"want_{i}")

                    st.markdown("---")
                    if st.button("✅ この内容でExcelを更新してダウンロード", type="primary"):
                        selected = [item for item in extracted if item.get("_include", True)]
                        if not selected:
                            st.warning("更新する企業を1社以上選択してください")
                        else:
                            updated_bytes = update_excel_with_requirements(selected, companies)
                            st.download_button(
                                label="📥 更新済みExcelをダウンロード",
                                data=updated_bytes,
                                file_name="求人確度出力マスターシート.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                            st.success(f"{len(selected)}社分を更新しました。ファイルを差し替えてください。")
